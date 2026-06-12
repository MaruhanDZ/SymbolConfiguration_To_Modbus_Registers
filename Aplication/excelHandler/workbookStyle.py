from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Estilos
header_fill = PatternFill("solid", fgColor="4472C4")  # Azul Excel
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center")



# Define o estilo do excel
titulo_fill = PatternFill("solid", fgColor="4472C4")
titulo_font = Font(color="FFFFFF", bold=True, size=12)
#header_fill = PatternFill("solid", fgColor="D9EAF7")
bold_font = Font(bold=True)
mono_font = Font(name="Consolas", size=10)


thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"))



def adjustWidth(ws, limite=200, additionalLen = 5):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)

        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )

        ws.column_dimensions[col_letter].width = min(max_length + additionalLen, limite)


# Faz a aba de informações do projeto
def configTab(projectTAB, symbolInfo, title):
    # Nome da aba
    projectTAB.title = title 

    linha = 1
    # Faz o cabeçalho
    projectTAB.cell(row=linha, column=1).value = "CONFIGURATION:"
    projectTAB.cell(row=linha, column=1).fill = titulo_fill
    projectTAB.cell(row=linha, column=1).font = titulo_font
    projectTAB.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
    linha += 1

    # imprime em cada linha uma das informações
    for chave, valor in symbolInfo.items():
        projectTAB.cell(row=linha, column=1).value = chave
        projectTAB.cell(row=linha, column=2).value = str(valor)
        projectTAB.cell(row=linha, column=1).font = bold_font

        for col in range(1, 3):
            cell = projectTAB.cell(row=linha, column=col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        linha += 1

    projectTAB.freeze_panes = "A2"
    projectTAB.auto_filter.ref = projectTAB.dimensions

    adjustWidth(projectTAB, limite=150, additionalLen=15)
    #return projectTAB

def treeTab(treeTAB, tree_str, title):

    # Nome da aba
    treeTAB.title = title

    linha = 1

    treeTAB.cell(row=linha, column=1).value = "Variable Tree:"
    treeTAB.cell(row=linha, column=1).fill = titulo_fill
    treeTAB.cell(row=linha, column=1).font = titulo_font
    linha += 1

    for tree_line in tree_str.splitlines():
        cell = treeTAB.cell(row=linha, column=1)
        cell.value = tree_line
        cell.font = mono_font
        cell.alignment = Alignment(wrap_text=False, vertical="center")
        linha += 1

    treeTAB.column_dimensions["A"].width = 180

    for row in range(1, linha + 1):
        treeTAB.row_dimensions[row].height = 18

    treeTAB.freeze_panes  = "A2"

    #return treeTAB


def createHeader(excelTab, headers):
    # adiciona o cabeçalho na ordem que foi passado
    excelTab.append(headers)

    # Formata cabeçalho
    for col in range(1, len(headers) + 1):
        cell = excelTab.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    # Congela a primeira linha
    excelTab.freeze_panes = "A2"


# adiciona uma variavel a aba do excel
def appendVar(excelTab, headers, mappings):
    # percorre a lista de variaveis mapeadas e converte para uma lista
    for mapping in mappings:
        linha = [mapping.get(campo, "") for campo in headers]
        excelTab.append(linha)

def convertToNumber(ws):
    for row in ws.iter_rows():
        for cell in row:
            valor = cell.value

            if isinstance(valor, str):
                texto = valor.strip()

                if texto == "":
                    continue

                try:
                    if "," in texto and "." not in texto:
                        numero = float(texto.replace(",", "."))
                    elif "." in texto:
                        numero = float(texto)
                    else:
                        numero = int(texto)

                    cell.value = numero

                except ValueError:
                    pass
 


