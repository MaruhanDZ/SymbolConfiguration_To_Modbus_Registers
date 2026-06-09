import xml.etree.ElementTree as ET
import pandas as pd
import math
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURAÇÃO
# ============================================================

XML_FILE = "testexml.PLC_AC500_V3.Application.xml"


# ============================================================
# LEITURA XML
# ============================================================

tree = ET.parse(XML_FILE)
root = tree.getroot()

ns = {
    "sc": "http://www.3s-software.com/schemas/Symbolconfiguration.xsd"
}

# ============================================================
# TIPOS
# ============================================================

types = {}

# ------------------------------------------------------------
# TypeSimple
# ------------------------------------------------------------

for t in root.findall(".//sc:TypeSimple", ns):

    types[t.attrib["name"]] = {
        "kind": "simple",
        "name": t.attrib["name"],
        "iecname": t.attrib.get("iecname", ""),
        "size": int(t.attrib.get("size", 0)),
        "typeclass": t.attrib.get("typeclass", ""),
        "bitoffset": int(t.attrib.get("bitoffset", 0))
        if "bitoffset" in t.attrib else None
    }

# ------------------------------------------------------------
# TypeArray
# ------------------------------------------------------------

for t in root.findall(".//sc:TypeArray", ns):

    dim = t.find("sc:ArrayDim", ns)

    types[t.attrib["name"]] = {
        "kind": "array",
        "name": t.attrib["name"],
        "iecname": t.attrib.get("iecname", ""),
        "size": int(t.attrib.get("size", 0)),
        "basetype": t.attrib["basetype"],
        "minrange": int(dim.attrib["minrange"]),
        "maxrange": int(dim.attrib["maxrange"])
    }

# ------------------------------------------------------------
# TypeUserDef
# ------------------------------------------------------------

for t in root.findall(".//sc:TypeUserDef", ns):

    elements = []

    for e in t.findall("sc:UserDefElement", ns):

        elements.append({
            "name": e.attrib["iecname"],
            "type": e.attrib["type"],
            "offset": int(e.attrib["byteoffset"])
        })

    types[t.attrib["name"]] = {
        "kind": "udt",
        "name": t.attrib["name"],
        "iecname": t.attrib.get("iecname", ""),
        "size": int(t.attrib.get("size", 0)),
        "elements": elements
    }

# ============================================================
# AUXILIARES
# ============================================================

rows = []


def add_variable(path,
                 offset,
                 type_name,
                 size,
                 typeclass,
                 bit=None):



    if bit is not None:
        reg_bit = bit
    else:
        reg_bit = ""

    rows.append({
        "Variavel": path,
        "Tipo": type_name,
        "Classe": typeclass,
        "Offset Byte": offset,
        "Tamanho Bytes": size,
       # "Registrador": register,
        "Bit": reg_bit
    })


# ============================================================
# EXPANSÃO RECURSIVA
# ============================================================

def expand_type(type_name, prefix, base_offset):

    t = types[type_name]

    # --------------------------------------------------------
    # SIMPLE
    # --------------------------------------------------------

    if t["kind"] == "simple":

        if t["typeclass"] == "Bit":

            add_variable(
                prefix,
                base_offset,
                t["iecname"],
                1,
                "Bit",
                bit=t["bitoffset"]
            )

        else:

            add_variable(
                prefix,
                base_offset,
                t["iecname"],
                t["size"],
                t["typeclass"]
            )

        return

    # --------------------------------------------------------
    # ARRAY
    # --------------------------------------------------------

    if t["kind"] == "array":

        min_idx = t["minrange"]
        max_idx = t["maxrange"]

        total_items = max_idx - min_idx + 1

        item_size = types[t["basetype"]]["size"]

        for i in range(total_items):

            offset = base_offset + i * item_size

            expand_type(
                t["basetype"],
                f"{prefix}[{i}]",
                offset
            )

        return

    # --------------------------------------------------------
    # UDT
    # --------------------------------------------------------

    if t["kind"] == "udt":

        for elem in t["elements"]:

            expand_type(
                elem["type"],
                f"{prefix}.{elem['name']}",
                base_offset + elem["offset"]
            )

        return


# ============================================================
# NODES PUBLICADOS
# ============================================================

for node in root.findall(".//sc:NodeList//sc:Node", ns):

    node_type = node.attrib.get("type")

    if not node_type:
        continue

    node_name = node.attrib["name"]

    print(f"Processando: {node_name}")

    expand_type(
        node_type,
        node_name,
        0
    )

# ============================================================
# DATAFRAME
# ============================================================

df = pd.DataFrame(rows)

df = df.sort_values(
    by=["Offset Byte", "Bit"]
)

# ============================================================
# RESUMO DOS UDTs
# ============================================================

udt_summary = []

for name, t in types.items():

    if t["kind"] == "udt":

        udt_summary.append({
            "UDT": t["iecname"],
            "Bytes": t["size"],
            "Registros Modbus": math.ceil(t["size"] / 2)
        })

df_summary = pd.DataFrame(udt_summary)

# ============================================================
# EXCEL
# ============================================================

output = "Mapa_Modbus.xlsx"

with pd.ExcelWriter(output,
                    engine="openpyxl") as writer:

    df_excel = df[['Variavel', 'Tipo', 'Classe', 'Tamanho Bytes', 'Bit']]
    df_excel['Size (Bytes)'] = df_excel['Tamanho Bytes'].where((df_excel['Bit'] == 0) | (df_excel['Bit'] == ''), 0)

    

    df_excel = df_excel[['Variavel', 'Tipo', 'Classe', 'Size (Bytes)', 'Bit']]
    df_excel['Byte Inicial'] = 0
    df_excel['Byte Final'] = 0
    df_excel['Registrador'] = 0

    df_excel.to_excel(
        writer,
        sheet_name="Mapa Modbus",
        index=False
    )
    ws = writer.sheets["Mapa Modbus"]
    ws['k1'] = 'Byte Inicial:'
    ws['l1'] = 0
    # primeira coluna byte final
    ws['g2'] = formula = '=L1+d2'
    # formula byte final
    ws['g3'] = formula = '=g2+d2'
    for row in range(3, ws.max_row + 1):
        ws[f'G{row}'] = f'=G{row-1}+D{row}'

    # formula byte inicial
    ws['F2'] = '=IF(D2=0,F1,G2-D2)'

    for row in range(3, ws.max_row + 1):
        ws[f'F{row}'] = f'=IF(D{row}=0,F{row-1},G{row}-D{row})'

    # Registrador final
    ws['H2'] = '=IF(ISEVEN(F2),F2/2,H1)'

    for row in range(3, ws.max_row + 1):
        ws[f'H{row}'] = f'=IF(ISEVEN(F{row}),F{row}/2,H{row-1})'

    # Cores
    verde = PatternFill(fill_type="solid", fgColor="92D050")
    amarelo = PatternFill(fill_type="solid", fgColor="FFFF00")

    # Formatar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # A:G verde
    for col in range(1, 8):  # A=1, G=7
        ws.cell(row=1, column=col).fill = verde

    # H amarelo
    ws["H1"].fill = amarelo

    # K:L verde
    for col in range(11, 13):  # K=11, L=12
        ws.cell(row=1, column=col).fill = verde

    # Ajustar largura das colunas ao conteúdo
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Adiciona uma margem para não ficar apertado
        ws.column_dimensions[column_letter].width = max_length + 2

    ws.column_dimensions['l'].width = 11  # Variável

    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['F'].hidden = True
    ws.column_dimensions['G'].hidden = True

print(f"Arquivo gerado: {output}")